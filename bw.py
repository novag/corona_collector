#!/usr/bin/env python3

import influxdb
import io
import json
import locale
import os
import openpyxl
import re
import requests
import sys
import traceback
from datetime import datetime
from lxml import html


DEBUG = False

STATE = 'Baden-Württemberg'
STATE_SHORT = 'BW'

BASE_PATH = os.path.dirname(os.path.realpath(__file__))
CONFIG_PATH = os.path.join(BASE_PATH, 'config.json')
POPULATION_PATH = os.path.join(BASE_PATH, 'population.json')

with open(CONFIG_PATH) as file:
    CONFIG = json.load(file)

with open(POPULATION_PATH, encoding='utf-8') as file:
    POPULATION = json.load(file)


def notify(msg):
    print(msg)

    if 'pushover' not in CONFIG:
        return

    requests.post(
        'https://api.pushover.net/1/messages.json',
        json={
            'token': CONFIG['pushover']['token'],
            'user': CONFIG['pushover']['user'],
            'title': 'Corona {}: Fehler!'.format(STATE_SHORT),
            'message': msg
        }
    )


class CoronaParser:
    def __init__(self, db, tree, doc_bytes):
        self.db = db
        self.tree = tree
        self.wb = openpyxl.load_workbook(doc_bytes)
        self.ws_infected = self.wb['Infizierte Coronavirus in BW']
        self.ws_death = self.wb['Todesfälle Coronavirus in BW']

        locale.setlocale(locale.LC_TIME, "de_DE.utf8")

    def _store(self, data):
        if DEBUG:
            print(data)
            return

        try:
            if not self.db.write_points(data):
                raise Exception('ERROR: CoronaParser: _store: false')
        except influxdb.exceptions.InfluxDBServerError as e:
            raise Exception('ERROR: CoronaParser: _store: {}'.format(e))
        except influxdb.exceptions.InfluxDBClientError as e:
            raise Exception('ERROR: CoronaParser: _store: {}'.format(e))

    def _parse_web_datetime(self):
        dt_text = self.tree.xpath('normalize-space(//figcaption/text())').split('(')[-1]
        date = datetime.strptime(dt_text, 'Stand: %d. %B %Y, %H:%M Uhr)')
        self.dt = date.strftime('%Y-%m-%dT%H:%M:%SZ')

        return date

    def _normalize_county(self, county):
        county = county.replace('Stuttgart', 'Stuttgart (Stadt)')
        county = county.replace(' (Stadtkreis)', ' (Stadt)')

        return county

    def _calculate_per_population(self, county, infected, per_population):
        try:
            if county.endswith('(Stadt)'):
                raw_county = county.replace(' (Stadt)', '')
                population = POPULATION['city'][STATE_SHORT][raw_county]
            else:
                population = POPULATION['county'][STATE_SHORT][county]

            return round(infected * per_population / population, 2)
        except:
            notify('{} not found in population database.'.format(county))

        return None

    def _calculate_p10k(self, county, infected):
        return self._calculate_per_population(county, infected, 10000)

    def _calculate_p100k(self, county, infected):
        return self._calculate_per_population(county, infected, 100000)

    def _calculate_state_per_population(self, infected, per_population):
        population = POPULATION['state'][STATE_SHORT]

        return round(infected * per_population / population, 2)

    def _calculate_state_p10k(self, infected):
        return self._calculate_state_per_population(infected, 10000)

    def _calculate_state_p100k(self, infected):
        return self._calculate_state_per_population(infected, 100000)

    def parse(self):
        dt_row = 0
        county_row = 0

        for row in self.ws_infected.iter_rows(min_col=self.ws_infected.min_column, min_row=self.ws_infected.min_row, max_row=self.ws_infected.max_row, max_col=self.ws_infected.min_column):
            cell_a = row[0]

            if not cell_a.value:
                continue

            if cell_a.value.startswith('Stadt-/Landkreis'):
                dt_row = cell_a.row + 1
                county_row = cell_a.row + 2
                break

        excel_dt = self.ws_infected['B'][dt_row - 1].value
        web_dt = self._parse_web_datetime()

        if excel_dt.date() != web_dt.date():
            raise Exception('WARN: Date mismatch: No datetime available yet. Skipping run...')

        data = []
        infected_sum = 0
        death_sum = 0
        for row in self.ws_infected.iter_rows(min_col=self.ws_infected.min_column, min_row=county_row, max_row=self.ws_infected.max_row, max_col=self.ws_infected.min_column + 1):
            county = self._normalize_county(row[0].value.strip())

            if not county:
                continue

            if county == 'Summe':
                break

            infected = row[1].value
            infected_sum += infected

            death = self.ws_death[row[1].coordinate].value
            death_sum += death

            data.append({
                'measurement': 'infected_de_state',
                'tags': {
                    'state': STATE,
                    'county': county
                },
                'time': self.dt,
                'fields': {
                    'count': infected,
                    'p10k': self._calculate_p10k(county, infected),
                    'p100k': self._calculate_p100k(county, infected),
                    'death': death
                }
            })

        data.append({
            'measurement': 'infected_de',
            'tags': {
                'state': STATE
            },
            'time': self.dt,
            'fields': {
                'count': infected_sum,
                'p10k': self._calculate_state_p10k(infected_sum),
                'p100k': self._calculate_state_p100k(infected_sum),
                'death': death_sum
            }
        })

        self._store(data)

        return self.dt


data_url = 'https://sozialministerium.baden-wuerttemberg.de/de/gesundheit-pflege/gesundheitsschutz/infektionsschutz-hygiene/informationen-zu-coronavirus/lage-in-baden-wuerttemberg/'
if len(sys.argv) == 2:
    data_url = sys.argv[1]

r_web = requests.get(data_url, headers={'User-Agent': CONFIG['user_agent']})
if not r_web.ok:
    print('ERROR: failed to fetch data, status code: {}'.format(r.status_code))
    sys.exit(1)

try:
    tree = html.fromstring(r_web.text)
    table_url = tree.xpath('//a[@class="link-download" and contains(@href, ".xlsx")]/@href')[0]
    table_url = 'https://sozialministerium.baden-wuerttemberg.de/{}'.format(table_url)
except Exception as e:
    traceback.print_exc()
    notify(str(e))
    sys.exit(1)

r_excel = requests.get(table_url, headers={'User-Agent': CONFIG['user_agent']})
if not r_excel.ok:
    print('ERROR: failed to fetch excel workbook, status code: {}'.format(r_excel.status_code))
    sys.exit(1)

if DEBUG:
    db_client = None
else:
    db_client = influxdb.InfluxDBClient(
        host=CONFIG['db']['host'],
        port=CONFIG['db']['port'],
        username=CONFIG['db']['username'],
        password=CONFIG['db']['password']
    )
    db_client.switch_database(CONFIG['db']['database'])

corona_parser = CoronaParser(db_client, tree, io.BytesIO(r_excel.content))

try:
    dt = corona_parser.parse()
    print('Data updated at {}'.format(dt))
except Exception as e:
    traceback.print_exc()
    notify(str(e))

if not DEBUG:
    db_client.close()
